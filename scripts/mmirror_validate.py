# -*- coding: utf-8 -*-
"""M-Spiegel: reproduziert alle Modelltabellen AUS data/raw mit exakt den
Schritten/Regeln, die in Power Query (M) kodiert werden (Positions-Spalten,
Jahr-Filter, Missing-Token=null, Komma/Punkt-Locale, AGS-Ableitung, Unpivot,
Name->Code-Mapping). Vergleicht das Ergebnis 1:1 mit data/clean, um die
M-Logik VOR dem Power-BI-Umbau zu verifizieren (kein PBI noetig).
"""
import csv, os, re
R=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW=os.path.join(R,"data","raw"); CLEAN=os.path.join(R,"data","clean")

MISS={"-","x",".","...","/",""}
def to_int(v):           # mirror: Text.Remove(" ."), leer/Sonderzeichen->null
    v=(v or "").strip()
    if v in MISS: return None
    return int(v.replace(".","").replace(" ",""))
def to_float(v):         # mirror: " ." entfernen, "," -> "." , en-US
    v=(v or "").strip()
    if v in MISS: return None
    return float(v.replace(".","").replace(" ","").replace(",","."))
def ebene(code):
    code=code.strip()
    if code=="DG": return "DE"
    if code.isdigit(): return {2:"BL",3:"RB",5:"KR"}.get(len(code),"?")
    return "?"
def blc(code): return "DG" if code=="DG" else code[:2]

def raw_year_rows(fn, yearpat, mincols):
    out=[]
    with open(os.path.join(RAW,fn),encoding="cp1252") as f:
        for line in f:
            p=[c.strip() for c in line.rstrip("\n").split(";")]
            if len(p)>=mincols and re.match(yearpat,p[0]): out.append(p)
    return out

# ---- Rebuild jede Tabelle aus RAW ----
def build_schule():
    out=[]
    for p in raw_year_rows("21111-01-03-4.csv", r"^2023$", 10):
        if ebene(p[1])=="?": continue
        out.append([2023,p[1],p[2],ebene(p[1]),blc(p[1]),p[3],
                    to_int(p[4]),to_int(p[5]),to_int(p[6]),to_int(p[7]),
                    to_int(p[8]) if len(p)>8 else None, to_int(p[9]) if len(p)>9 else None])
    return ["jahr","region_code","region","ebene","bundesland_code","schulart","schulen","schueler_insg","schueler_w","schueler_ausl","klasse_7","jahrgang_11"], out

def build_arbeitsmarkt():
    out=[]
    for p in raw_year_rows("13211-02-05-4.csv", r"^2025$", 16):
        if ebene(p[1])=="?": continue
        out.append([2025,p[1],p[2],ebene(p[1]),blc(p[1]),
                    to_int(p[3]),to_int(p[4]),to_int(p[6]),to_int(p[7]),
                    to_float(p[10]),to_float(p[15])])
    return ["jahr","region_code","region","ebene","bundesland_code","arbeitslose_insg","arbeitslose_ausl","arbeitslose_15_20","arbeitslose_15_25","alq_insg","jugend_alq_15_25"], out

def build_bev():
    out=[]
    with open(os.path.join(RAW,"12411-02-03-4.csv"),encoding="cp1252") as f:
        for line in f:
            p=[c.strip() for c in line.rstrip("\n").split(";")]
            if len(p)<7: continue
            m=re.match(r"^31\.12\.(\d{4})$",p[0])
            if not m: continue
            jahr=int(m.group(1))
            if jahr not in (2023,2024): continue
            if ebene(p[1])=="?": continue
            out.append([jahr,p[1],p[2],ebene(p[1]),blc(p[1]),p[3],to_int(p[4]),to_int(p[5]),to_int(p[6])])
    return ["jahr","region_code","region","ebene","bundesland_code","altersgruppe","insgesamt","maennlich","weiblich"], out

def build_beruflich():
    out=[]
    for p in raw_year_rows("21121-02-02-4.csv", r"^2023$", 15):
        if ebene(p[1])=="?": continue
        out.append([2023,p[1],p[2],ebene(p[1]),blc(p[1]),
                    to_int(p[3]),to_int(p[5]),to_int(p[7]),to_int(p[9]),to_int(p[11])])
    return ["jahr","region_code","region","ebene","bundesland_code","insgesamt","mit_hauptschulabschluss","mit_mittlerem_abschluss","fachhochschulreife","allg_hochschulreife"], out

def build_einkommen():
    out=[]
    for p in raw_year_rows("82411-01-03-4.csv", r"^2021$", 5):
        if ebene(p[1])=="?": continue          # Sub-Kreis-VGRdL-Codes (8-stellig) raus
        v=to_int(p[4])
        if v is None: continue                 # nur Regionen mit vorhandenem Einkommenswert (wie Modell-M)
        out.append([p[1],2021,p[2],v])
    return ["region_code","jahr","region","einkommen_je_ew"], out

# abgaenge kreis (wide->long) -> intermediate (fuer fact_abgaenge + dim_region)
COLS=[("insgesamt",3,4),("ohne_hauptschulabschluss",5,6),("mit_hauptschulabschluss",7,8),
      ("mittlerer_abschluss",9,10),("dar_schul_teil_fhr",11,12),("fachhochschulreife",13,14),
      ("allgemeine_hochschulreife",15,16)]
def build_abg_kreis():
    recs=[]
    for p in raw_year_rows("21111-02-06-4.csv", r"^2023$", 17):
        if ebene(p[1])=="?": continue
        rec={"jahr":2023,"region_code":p[1],"region":p[2],"ebene":ebene(p[1]),"bundesland_code":blc(p[1])}
        for name,iI,iW in COLS: rec[name]=to_int(p[iI]); rec[name+"_w"]=to_int(p[iW])
        recs.append(rec)
    return recs

KEY={"ohne_hauptschulabschluss":"ohne_hauptschulabschluss","mit_hauptschulabschluss":"mit_hauptschulabschluss",
     "mittlerer_abschluss":"mittlerer_abschluss","fachhochschulreife":"fachhochschulreife",
     "allgemeine_hochschulreife":"allgemeine_hochschulreife","hauptschulabschluss":"mit_hauptschulabschluss",
     "mittlerer abschluss":"mittlerer_abschluss","allgemeine hochschulreife":"allgemeine_hochschulreife",
     "ohne hauptschulabschluss":"ohne_hauptschulabschluss","mit hauptschulabschluss":"mit_hauptschulabschluss"}

def build_statbericht_2022():
    """csv-21111-12 aus statbericht_2022-23.xlsx, nur 2022 BL-Ebene."""
    import openpyxl,warnings; warnings.filterwarnings("ignore")
    wb=openpyxl.load_workbook(os.path.join(RAW,"statbericht_allgbild_2022-23.xlsx"),read_only=True,data_only=True)
    ws=wb["csv-21111-12"]; data=list(ws.iter_rows(values_only=True))
    header=[("" if c is None else str(c)).strip() for c in data[0]]
    def col(name,exclude=None):
        for i,c in enumerate(header):
            if name.lower() in c.lower() and (exclude is None or exclude.lower() not in c.lower()): return i
        return None
    iBL,iSA,iST,iKL,iA2,iAB,iGE=col("Bundesland"),col("Schulart"),col("Status"),col("Klassenstufe"),col("Abschluss2"),col("Abschluss"),col("Geschlecht")
    iVAL=col("Absolvierende_und_Abgehende_Anzahl",exclude="auslaend")
    def ins(x): return ("" if x is None else str(x)).strip().lower()=="insgesamt"
    rows=[]
    for r in data[1:]:
        r=[("" if c is None else str(c)).strip() for c in r]
        if not(ins(r[iSA]) and ins(r[iST]) and ins(r[iKL]) and ins(r[iA2])): continue
        if not(ins(r[iGE]) or r[iGE] in ("männlich","weiblich")): continue
        rows.append({"bundesland":r[iBL],"abschluss":r[iAB],"geschlecht":r[iGE],"anzahl":to_int(r[iVAL])})
    return rows

def build_dim_region(abg_kreis):
    OST={"12","13","14","15","16"}; reg={}
    for r in abg_kreis:
        c=r["region_code"]
        if c in reg: continue
        name=r["region"]; e=r["ebene"]; bl=r["bundesland_code"]
        stadt="Stadt" if ("kreisfreie Stadt" in name or "Stadtkreis" in name) else ("Land" if e=="KR" else "")
        ow="Berlin" if bl=="11" else ("Ost" if bl in OST else ("West" if e in ("KR","BL","RB") else ""))
        reg[c]=[c,name,e,bl,stadt,ow]
    return ["region_code","region","ebene","bundesland_code","stadt_land","ost_west"], list(reg.values())

def build_fact_abgaenge(abg_kreis, dimreg_rows, statb):
    n2c={r[1]:r[0] for r in dimreg_rows if r[2]=="BL"}; n2c["Deutschland"]="DG"; n2c["Zusammen"]="DG"
    out=[]
    for r in abg_kreis:
        for name,_,_ in COLS:
            k=KEY.get(name)
            if not k: continue
            insg=r[name]; w=r[name+"_w"]
            out.append([r["region_code"],2023,k,"insgesamt", insg if insg is not None else ""])
            out.append([r["region_code"],2023,k,"weiblich", w if w is not None else ""])
            m=(insg-w) if (insg is not None and w is not None) else None
            out.append([r["region_code"],2023,k,"maennlich", m if m is not None else ""])
    for r in statb:
        k=KEY.get(r["abschluss"].strip().lower())
        if not k: continue
        code=n2c.get(r["bundesland"].strip())
        if not code: continue
        g=r["geschlecht"].strip().lower().replace("ä","ae")
        a=r["anzahl"]
        out.append([code,2022,k,g, a if a is not None else ""])
    return ["region_code","jahr","abschluss_key","geschlecht","anzahl"], out

def build_ausgaben_xlsx(sheet):
    import openpyxl,warnings; warnings.filterwarnings("ignore")
    wb=openpyxl.load_workbook(os.path.join(RAW,"21711_ausgaben_je_schueler_2024.xlsx"),read_only=True,data_only=True)
    ws=wb[sheet]; data=list(ws.iter_rows(values_only=True))
    header=[("" if c is None else str(c)).strip() for c in data[0]]
    H={c:i for i,c in enumerate(header)}
    rows=[]
    for r in data[1:]:
        r=[("" if c is None else str(c)).strip() for c in r]
        if len(r)<=H["Ausgaben_je_Schueler"]: continue
        rows.append(r)
    return H, rows

def build_ausgaben_je_schueler(dimreg_rows):
    n2c={r[1]:r[0] for r in dimreg_rows if r[2]=="BL"}; n2c["Deutschland"]="DG"
    H,rows=build_ausgaben_xlsx("csv-21711-b01")
    gi,ji,vi=H["Gebiet"],H["Jahr"],H["Ausgaben_je_Schueler"]
    out=[]
    for r in rows:
        j=to_int(r[ji]); v=to_int(r[vi]); g=r[gi].strip()
        if j is None or v is None: continue
        out.append([n2c.get(g,""),g,j,"Alle Schularten",v])
    return ["region_code","bundesland","jahr","schulart","ausgaben_je_schueler"], out

def build_ausgaben_schulart(dimreg_rows):
    n2c={r[1]:r[0] for r in dimreg_rows if r[2]=="BL"}; n2c["Deutschland"]="DG"
    H,rows=build_ausgaben_xlsx("csv-21711-02")
    gi,si,ji,vi=H["Gebiet"],H["Schulart"],H["Jahr"],H["Ausgaben_je_Schueler"]
    def isint(x):
        try: int(x); return True
        except: return False
    out=[]
    for r in rows:
        g,s,j,v=r[gi].strip(),r[si].strip(),r[ji].strip(),r[vi].strip()
        if isint(j) and isint(v): out.append([n2c.get(g,""),g,s,int(j),int(v)])
    return ["region_code","bundesland","schulart","jahr","ausgaben_je_schueler"], out

# ---- Vergleich ----
def norm(v):
    if v is None: return ""
    return str(v)
def load_clean(fn):
    with open(os.path.join(CLEAN,fn),encoding="utf-8") as f:
        rows=list(csv.reader(f,delimiter=";"))
    return rows[0], rows[1:]
def compare(name, built_hdr, built_rows, clean_fn, key_sort=True):
    ch,cr=load_clean(clean_fn)
    bstr=[[norm(x) for x in r] for r in built_rows]
    ok_hdr = built_hdr==ch
    # als Menge (Reihenfolge kann in M abweichen) vergleichen
    bs=sorted(tuple(r) for r in bstr); cs=sorted(tuple(r) for r in cr)
    same = bs==cs
    only_b=[r for r in bs if r not in set(cs)][:3]
    only_c=[r for r in cs if r not in set(bs)][:3]
    print(f"[{'OK ' if (ok_hdr and same) else 'XX '}] {name:28s} built={len(bstr):5d} clean={len(cr):5d} header={'ok' if ok_hdr else ch}")
    if not same:
        print("     nur_in_M(3):",only_b)
        print("     nur_in_clean(3):",only_c)
    return ok_hdr and same

abg_kreis=build_abg_kreis()
dh,dr=build_dim_region(abg_kreis)
statb=build_statbericht_2022()
results=[]
results.append(compare("fact_schule_2023",*build_schule(),"fact_schule_2023.csv"))
results.append(compare("fact_arbeitsmarkt_2025",*build_arbeitsmarkt(),"fact_arbeitsmarkt_2025.csv"))
results.append(compare("fact_bevoelkerung_2023_2024",*build_bev(),"fact_bevoelkerung_2023_2024.csv"))
results.append(compare("fact_abgaenge_beruflich_2023",*build_beruflich(),"fact_abgaenge_beruflich_2023.csv"))
results.append(compare("fact_einkommen_kreis",*build_einkommen(),"fact_einkommen_kreis.csv"))
results.append(compare("dim_region",dh,dr,"dim_region.csv"))
results.append(compare("fact_abgaenge",*build_fact_abgaenge(abg_kreis,dr,statb),"fact_abgaenge.csv"))
results.append(compare("fact_ausgaben_je_schueler",*build_ausgaben_je_schueler(dr),"fact_ausgaben_je_schueler.csv"))
results.append(compare("fact_ausgaben_schulart",*build_ausgaben_schulart(dr),"fact_ausgaben_schulart.csv"))
print(f"\n{sum(results)}/{len(results)} Tabellen 1:1 reproduziert aus data/raw")
